#!/usr/bin/env python3
# -*- coding: utf-8 -*-
'''
程序用于excel表格的复制汇总，可以选择是workbook中某张sheet全部复制，也可以是sheet中某个列的复制。
author:WYF
DATE:2019/7/29
'''

import os
import openpyxl


# 寻址
def get_file_name(folder_name):
    folder_path = os.path.join(path, folder_name)
    folder_dir_name = os.listdir(folder_path)
    file_name_path = list(map(lambda x: os.path.join(folder_path, x), folder_dir_name))
    return file_name_path


# 全表复制
def copy_all(sheet1, sheet2, max_row):
    for row in range(1, sheet1.max_row + 1):
        for col in range(1, sheet1.max_column + 1):
            sheet2.cell(max_row + row, col).value = sheet1.cell(row, col).value


# 列复制
def copy_col(sheet1, sheet2, col_number, max_row):
    for row in range(1, sheet1.max_row + 1):
        for col in col_number:
            sheet2.cell(max_row + row, col_number.index(col)+1).value = sheet1.cell(row, col).value


# 行复制
def copy_row(sheet1, sheet2, row_number, max_row):
    for row in row_number:
        for col in range(1, sheet1.max_column + 1):
            sheet2.cell(max_row + row_number.index(row)+1, col).value = sheet1.cell(row, col).value


# 表格的打开、写入、关闭
def deal_execl(file_name, sheet_name, func_number):
    wb = openpyxl.load_workbook(file_name)
    ws = wb[sheet_name]
    if os.path.exists(file_hz):
        wb_hz = openpyxl.load_workbook(file_hz)
        ws_hz = wb_hz.active
        row_max_number = ws_hz.max_row
    else:
        wb_hz = openpyxl.Workbook()
        ws_hz = wb_hz.active
        row_max_number = ws_hz.max_row - 1

    if func_number == 1:
        copy_all(ws, ws_hz, row_max_number)
    elif func_number == 2:
        copy_col(ws, ws_hz, Col_number, row_max_number)
    elif func_number == 3:
        copy_row(ws, ws_hz, Row_number, row_max_number)
    wb_hz.save(file_hz)
    wb_hz.close()
    wb.close()


# 运行程序
def run(name_list):
    for mz in name_list:
        print("进行中：{0}/{1}......".format(name_list.index(mz) + 1, len(name_list)))
        deal_execl(mz, Sheet_name, choice)
        print('Done.')


# 主程序
print(
    "此程序用于excel表格的复制汇总。\n请将包含要处理的表格的文件夹放至桌面。程序最终会在桌面生成汇总表格'汇总.xlsx'。\n注：一个excel文件是一个工作簿，其中包含多个表格。")
Folder_name = input("请输入文件夹名称：")
Sheet_name = input("请输入指定的表的名称[Sheet1]：")
if Sheet_name == '':
    Sheet_name = 'Sheet1'
output_file = input("请输入保存文件的名称：")

path = os.path.join(os.path.expanduser("~"), "Desktop")
file_hz = os.path.join(path, output_file)

Name_list = get_file_name(Folder_name)
print("请选择功能：\n\t1、表格复制\n\t2、列复制\n\t3、行复制")
choice = int(input())
if choice == 1:
    print("您选择的是：表格复制。")
    run(Name_list)
elif choice == 2:
    print("您选择的是：列复制。")
    print("请输入列号,如：ABC")
    col_alp = input('')
    Col_number = list(map(lambda x: openpyxl.utils.column_index_from_string(x), col_alp))
    Col_number.sort()
    run(Name_list)
elif choice == 3:
    print("您选择的是：行复制。")
    print("请输入行号,如：123")
    row_str = input('')
    Row_number = list(map(lambda x: int(x), row_str))
    Row_number.sort()
    run(Name_list)
else:
    print("错误输入。")

# 程序结束
print('Mession completed!.')
os.system('pause')
