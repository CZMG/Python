# -*- coding: utf-8 -*-
"""
Created on Tue Jul 23 21:26:46 2019

@author: WYF

"""

import os
import openpyxl
from  Decorator_list import running_time


def pd(sheet_name, workbook):
    if sheet_name == '':
        worksheet = workbook.active
    else:
        worksheet = workbook[sheet_name]
    return worksheet


# deal with files
@running_time
def deal_files(file1, file2, sheet1, sheet2, col):
    wb1 = openpyxl.load_workbook(file1)
    ws1 = pd(sheet1, wb1)
    wb2 = openpyxl.load_workbook(file2)
    ws2 = pd(sheet2, wb2)
    mark_red(ws1, judge(get_assign_col_value(col, ws1), ws2))
    wb1.save(file1)
    wb1.close()
    wb2.save(file2)
    wb2.close()


# 获取指定列的值
def get_assign_col_value(col_name, sheet):
    col_num = openpyxl.utils.column_index_from_string(col_name)
    content = [sheet.cell(i, col_num).value for i in range(1, len(sheet[col_name + '0']) + 1)]
    return content


# 判断原文件指定列的值是否在对比文件中。若在，则记录下行号。
def judge(s_content, sheet):
    result = []
    for i in range(1, sheet.max_row + 1):
        d_row_content = [sheet.cell(i, j).value for j in range(1, sheet.max_column + 1)]
        result += [s_content.index(k) + 1 for k in s_content if k in d_row_content]
    return set(result)


# 标红
def mark_red(sheet, row_list):
    ft = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    for row in row_list:
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row, col).font = ft


if __name__ == '__main__':
    print("此程序用于对比excel文件A和B。判断A中指定列的值是否在B中，若在，则将存在的行标红。")
    name_s = input("原文件：")
    sheet_s = input("表名[默认第一章表]:")
    col_user = input("列号：")
    name_d = input("对比文件：")
    sheet_d = input("表名[默认第一章表]：")
    file_path = os.path.join(os.path.expanduser('~'), 'Desktop')
    file_s = os.path.join(file_path, name_s)
    file_d = os.path.join(file_path, name_d)
    deal_files(file_s, file_d, sheet_s, sheet_d, col_user)
    print("Mession Done.")
    os.system('pause')
