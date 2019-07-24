# -*- coding: utf-8 -*-
"""
Created on Tue Jul 23 21:26:46 2019

@author: WYF

"""

import openpyxl
import os

# load file
print("此程序用于对比excel文件A和B。判断A中指定列的值是否在B中，若在，则将存在的行标红，并且在A中新建两个表格，分别为保存存在的数据和不存在的数据。")
name_s = input("原文件：")
col_user = input("列号：")
name_d = input("对比文件：")
file_path = os.path.join(os.path.expanduser('~'),'Desktop')
file_s = os.path.join(file_path,name_s)
file_d = os.path.join(file_path,name_d)

wb_s = openpyxl.load_workbook(file_s)
wb_d = openpyxl.load_workbook(file_d)
ws_s_1 = wb_s.active
ws_d = wb_d.active
ws_s_2 = wb_s.create_sheet("已存在")
ws_s_3 = wb_s.create_sheet("不存在")

# 获取指定列的值
number_max = len(ws_s_1[col_user+"0"])
l_s = col_user + "1"
l_e = col_user + str(number_max)

s_content = []
for col in ws_s_1[l_s:l_e]:
    for cell in col:
        s_content.append(cell.value)
    
# 判断原文件指定列的值是否在对比文件中。若在，则记录下行号。
d_content = []
for row in ws_d.rows:
    for cell_d in row:
        d_content.append(cell_d.value)

row_number = []
for i in range(len(s_content)):
    if s_content[i] in d_content:
        row_number.append(i+1)

# 标红，并将已存在和不存在的条目分别写入新的表中。
ft = openpyxl.styles.Font(color = openpyxl.styles.colors.RED)
for j in row_number:
    ws_s_1[col_user+str(j)].font = ft
    
for j in range(len(row_number)):
    for k in range(ws_s_1.max_column):
        ws_s_1.cell(row_number[j],k+1).font = ft    
        ws_s_2.cell(j+1,k+1).value = ws_s_1.cell(row_number[j],k+1).value

no_number = []
for l in range(1,number_max+1):
    if l not in row_number:
        no_number.append(l)

for m in range(len(no_number)):
    for n in range(ws_s_1.max_column):
        ws_s_3.cell(m+1,n+1).value = ws_s_1.cell(no_number[m],n+1).value
    
wb_s.save(file_s)
wb_d.save(file_d)
