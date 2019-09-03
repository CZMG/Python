# -*- coding: utf-8 -*-
'''
用于读取excel中的IP地址那一列，将IP进行升序排序。
'''
import os
import openpyxl


def get_file(name):
    path_desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
    file_path = os.path.join(path_desktop, name)
    return file_path


def deal_file(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    ip_list = [ws.cell(i, 1).value for i in range(1, len(ws['A0']) + 1)]
    new_ip = sort_ip(ip_list)
    for j in range(1, len(ws['A0']) + 1):
        ws.cell(j, 1).value = new_ip[j - 1]
    wb.save(file)
    return print("排序已完成。")


def sort_ip(list):
    ip_new_list = sorted(list,key = lambda x: ( int(x.split('.')[0]), int(x.split('.')[1]), int(x.split('.')[2]), int(x.split('.')[3])))
    return ip_new_list


if __name__ == "__main__":
    name_excel = input('请输入文件名：\n>>>')
    deal_file(get_file(name_excel))
    os.system('pause')

