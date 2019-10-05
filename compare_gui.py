#!/usr/bin/env python3
# -*- encoding: utf-8 -*-
'''
@File    :   compare_ui.py
@Time    :   2019/10/04 17:52:35
@Author  :   WYF 
@Version :   1.0
@Contact :   1007653544@qq.com
'''

# here put the import lib
import tkinter as tk
from tkinter import filedialog
import time
import openpyxl
from functools import wraps

def running_time(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        start = time.time()
        res = func(*args, **kwargs)
        end = time.time()
        txt.insert('end',"Running time is: %.2s s." % (end - start))
        return res
    return wrapper

def pd(sheet_name, workbook):
    if sheet_name == '':
        worksheet = workbook.active
    else:
        worksheet = workbook[sheet_name]
    return worksheet


@running_time
# deal with files
def deal_files(file1, file2, sheet1, sheet2, col):
    wb1 = openpyxl.load_workbook(file1)
    ws1 = pd(sheet1, wb1)
    wb2 = openpyxl.load_workbook(file2)
    ws2 = pd(sheet2, wb2)
    mark_red(ws1, judge(get_assign_col_value(col, ws1), ws2))
    wb1.save(file1)
    wb1.close()
    # wb2.save(file2)
    wb2.close()


# 获取指定列的值
def get_assign_col_value(col_name, sheet):
    col_num = openpyxl.utils.column_index_from_string(col_name)
    content = [sheet.cell(i, col_num).value for i in range(1, len(sheet[col_name + '0']) + 1)]
    return content


# 判断原文件指定列的值是否在对比文件中。若在，则记录下行号。
def judge(s_content, sheet):
    result = []
    for i in range(1, sheet.max_row + 1):
        txt.insert('end',"The {} line starts... \n".format(i))
        d_row_content = [sheet.cell(i, j).value for j in range(1, sheet.max_column + 1)]
        result += [s_content.index(k)+1 for k in s_content if k in d_row_content]
    return set(result)


# 标红
def mark_red(sheet, row_list):
    global data_deal
    ft = openpyxl.styles.Font(color=openpyxl.styles.colors.RED)
    data_deal = len(row_list)
    for row in row_list:
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row, col).font = ft


def button_choice_file(name, place_x, place_y):
    name_dict[name]=[]
    var = tk.StringVar()
    tk.Label(window, textvariable=var, bg="gray", font=("楷体", 14), height=1, width=60).place(relx=place_x, y=place_y)
    def open_file():
        name_path = filedialog.askopenfilename()
        if name_path != '':
            name_dict[name].append(name_path)
            var.set(name_path)
        else:
            var.set("请选择")
    tk.Button(window, text=name, command=open_file, relief="raised",height=2, width=20).place(relx=0.01, y=place_y)
    tk.Label(window, text = "表名[默认第一张表]:",font=("楷体", 12),height=1).place(relx=place_x,y=place_y+30)


def get_sheet_col(name, place_x, place_y):
    ent = tk.Entry(window,background='white',width=20)
    ent.place(relx=0.4,y=place_y+30)
    if name == "原文件":
        tk.Label(window,text="列号:",font=("楷体",12),height=1).place(relx=0.6,y=place_y+30)  
        ent2 = tk.Entry(window,background='white',width=5)
        ent2.place(relx=0.66,y=place_y+30)
    def get_var():
        name_dict[name].append(ent.get())
        if name == "原文件":
            name_dict[name].append(ent2.get())
        sure_button.config(state='disabled')
    sure_button = tk.Button(window,text="确认",command=get_var,relief="raised",height=1,state='normal')
    sure_button.place(relx=0.8,y=place_y+27)


def button_start():
    file_s = name_dict["原文件"][0]
    sheet_s = name_dict["原文件"][1]
    col_user = name_dict["原文件"][2]
    file_d = name_dict["对比文件"][0]
    sheet_d = name_dict["对比文件"][1]
    txt.insert('end',"Start to run.\n")
    deal_files(file_s, file_d, sheet_s, sheet_d, col_user)
    txt.insert('end','\n共处理{}行数据。\n'.format(data_deal))
    txt.insert('end', "The mession has done.\n")
    txt.yview_moveto(1)
    txt.update()

if __name__ == '__main__':
    window = tk.Tk()
    window.geometry("800x600")
    window.title("Compare")
    name_dict = {}
    data_deal = 0
    button_choice_file("原文件", 0.2, 20)
    get_sheet_col("原文件",0.2,20)
    button_choice_file("对比文件", 0.2, 100)
    get_sheet_col("对比文件",0.2,100)
    tk.Button(window,text="运行",relief="raised",height=2,width=20,command=button_start).place(relx=0.01,y=180)
    tk.Button(window,text="退出",relief="raised",height=2,width=20,command=window.quit).place(relx=0.01,y=260)
    sb = tk.Scrollbar(window)
    sb.pack(side='right',fill='y')
    txt = tk.Text(window,  background="white", font=("Segoe Script", 12), relief="raised", height=10,width=50,yscrollcommand=sb.set)
    txt.place(relx=0.2,y=180)
    sb.config(command=txt.yview)
    tk.Label(window, text="Excel指定数据查找（标红）", height=1, font=("黑体", 18), anchor='w').place(relx=0.1, rely=0.85)
    tk.Label(window, text="吴叶峰\nwuyefeng@zbiti.com\n2019/10/3", font=("华文行楷", 14),height=3, justify='right', anchor='e').place(relx=0.78, rely=0.85)
    window.mainloop()