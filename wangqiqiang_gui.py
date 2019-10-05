#!/usr/bin/env python3
# -*- encoding: utf-8 -*-
'''
@File    :   wangqiqiang.py
@Time    :   2019/10/01 19:45:16
@Author  :   WYF 
@Version :   3.0
@Contact :   1007653544@qq.com
'''

# here put the import lib
import os
import openpyxl
import time
import ipaddress
import tkinter as tk
from tkinter import filedialog
from functools import wraps


def running_time(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        start = time.time()
        res = func(*args, **kwargs)
        end = time.time()
        result.append("Running time is: %.2s s." % (end - start))
        return res
    return wrapper


def model_button(name, i):
    var = tk.StringVar()
    tk.Label(window, textvariable=var, bg="gray", font=(
        "楷体", 14), height=2, width=60).place(relx=0.2, y=20+60*(i-1))

    def print_thing():
        content = filedialog.askopenfilename()
        if content != '':
            name_dict[name] = content
            var.set(content)
        else:
            var.set("请选择")
    tk.Button(window, text=name, command=print_thing, relief="raised",
              height=2, width=20).place(relx=0.01, y=20+60*(i-1))


def run():
    global org_name
    file_hl = name_dict["互联地址段及反馈表"]
    file_ds = name_dict["地市表"]
    file_city = name_dict["地市信息表"]
    dishi = file_ds.split("/")[-1].split(".")[0]
    org_name = dishi + '电信'
    file_result = os.path.join(os.path.join(os.path.expanduser("~"), "Desktop"),  dishi + "result.xlsx")
    deal_excel(file_result, file_hl, file_ds, file_city)
    for i in result:
        txt.insert('end', i+'\n')
    txt.insert('end', "共{}行数据。\n".format(number))
    txt.insert('end', "The mession has done.\n")
    txt.yview_moveto(1)


def hl_excel(sheet, sheet_hl, sheet_ds):
    '''
    将互联及地址段反馈表和地市表填入最终表格。
    '''
    global number
    number = sheet_hl.max_row - 1
    for i in range(2, sheet_hl.max_row+1):
        result.append("互联、地市表：第{}行.".format(i))
        ip_str = sheet_hl.cell(i, 1).value
        ip = ipaddress.ip_interface(ip_str)
        ip_mask = ip.network.netmask
        ip_start = str(ip.ip)
        ip_compare = '.'.join(ip_start.split(".")[:-1])
        sheet.cell(i-1, 3).value = ip_start
        sheet.cell(i-1, 4).value = str(ip.network[-1])
        sheet.cell(i-1, 19).value = ip_compare + '.1'
        sheet.cell(i-1, 12).value = sheet_hl.cell(i, 2).value
        if sheet_hl.cell(i, 2).value == '预留':
            sheet.cell(i-1, 8).value = '其他自用地址'
        else:
            sheet.cell(i-1, 8).value = '设备和互联地址'
        for j in range(1, sheet_ds.max_row + 1):
            if sheet_ds.cell(j, 1).value == ip_compare + '.0':
                sheet.cell(i-1, 9).value = sheet_ds.cell(j, 9).value
                sheet.cell(i-1, 10).value = sheet_ds.cell(j, 10).value
                sheet.cell(i-1, 11).value = sheet_ds.cell(j, 11).value
        sheet.cell(i-1, 1).value = '新增'
        sheet.cell(i-1, 2).value = org_name
        sheet.cell(i-1, 14).value = time.strftime('%Y-%m-%d', time.localtime())


def city_excel(sheet, sheet_city):
    '''
    将地市信息表的信息填入最终表格。
    '''
    row_num = 0
    col_list = [5, 6, 7, 13, 16, 17, 18, 20, 21, 22, 23, 24, 25, 26, 27, 28]
    for i in range(1, sheet_city.max_row + 1):
        if sheet_city.cell(i, 2).value == org_name:
            row_num = i
            break
    for j in range(1, number+1):
        result.append("地市信息表：第{}行.".format(j))
        for k in col_list:
            sheet.cell(j, k).value = sheet_city.cell(row_num, k).value


@running_time
def deal_excel(result_name, hl_name, ds_name, city_name):
    wb_hl = openpyxl.load_workbook(hl_name)
    ws_hl = wb_hl["反馈表"]
    wb_ds = openpyxl.load_workbook(ds_name)
    ws_ds = wb_ds["Sheet1"]
    wb_city = openpyxl.load_workbook(city_name)
    ws_city = wb_city.active
    wb_result = openpyxl.Workbook()
    ws_result = wb_result.active
    hl_excel(ws_result, ws_hl, ws_ds)
    city_excel(ws_result, ws_city)
    wb_result.save(result_name)
    wb_result.close()
    wb_city.close()
    wb_ds.close()
    wb_hl.close()


if __name__ == '__main__':
    window = tk.Tk()
    window.title("地市资产处理")
    window.geometry("800x600")
    name_dict = {}
    model_button("互联地址段及反馈表", 1)
    model_button("地市表", 2)
    model_button("地市信息表", 3)
    result = []
    number = 0
    org_name = ''
    tk.Button(window, text="运行", command=run, relief="raised",
              height=2, width=20).place(relx=0.01, y=200)
    txt = tk.Text(window,  background="white", font=(
        "Segoe Script", 12), relief="raised", height=10, width=50)
    txt.place(relx=0.2, y=200)
    tk.Button(window, text="退出", height=2, width=20,
              command=window.quit).place(relx=0.01, y=260)
    tk.Label(window, text="江苏省各地市资产信息处理", height=1, font=(
        "黑体", 18), anchor='w').place(relx=0.1, rely=0.85)
    tk.Label(window, text="吴叶峰\nwuyefeng@zbiti.com\n2019/10/3", font=("华文行楷", 14),
             height=3, justify='right', anchor='e').place(relx=0.78, rely=0.85)
    window.mainloop()
